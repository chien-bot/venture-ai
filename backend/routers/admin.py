"""
routers/admin.py
────────────────────────────────────────────────────────────────
教务管理端 API — Admin Portal

功能：
1. 用户与权限管理：账号增删改查、角色分配
2. 全局数据监控：全校项目宏观统计、整体健康度、高频业务漏洞看板
3. RBAC 越权拦截：非 admin 角色访问返回 403
"""
import io
import json
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from services.database import (
    get_user_by_token,
    get_all_projects,
    get_conn,
    get_teacher_class_ids,
    set_teacher_class_ids,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── RBAC Guard ─────────────────────────────────────────────────
def _require_admin(request: Request) -> dict:
    """验证请求者是否为 admin 角色，否则返回 403。"""
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "").strip()
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token 无效")
    if user.get("role") != "admin":
        raise HTTPException(
            status_code=403,
            detail="Unauthorized Access Attempt: 仅管理员可访问此接口",
        )
    return user


# ── 1. 用户管理 ─────────────────────────────────────────────────

@router.get("/users")
def list_users(request: Request):
    """获取所有用户列表（仅管理员）。"""
    _require_admin(request)
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT user_id, username, role, display_name, class_id, created_at FROM users ORDER BY created_at DESC"
        ).fetchall()
    return {"users": [dict(r) for r in rows], "total": len(rows)}


class UpdateRoleRequest(BaseModel):
    role: str  # student | teacher | admin


@router.put("/users/{user_id}/role")
def update_user_role(user_id: str, req: UpdateRoleRequest, request: Request):
    """修改用户角色（仅管理员）。"""
    _require_admin(request)
    if req.role not in ("student", "teacher", "admin"):
        raise HTTPException(status_code=400, detail="角色必须为 student/teacher/admin")
    with get_conn() as conn:
        result = conn.execute(
            "UPDATE users SET role=? WHERE user_id=?", (req.role, user_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="用户不存在")
    return {"ok": True, "user_id": user_id, "new_role": req.role}


class UpdateClassRequest(BaseModel):
    class_id: str


@router.put("/users/{user_id}/class")
def update_user_class(user_id: str, req: UpdateClassRequest, request: Request):
    """修改用户班级（仅管理员）。"""
    _require_admin(request)
    with get_conn() as conn:
        result = conn.execute(
            "UPDATE users SET class_id=? WHERE user_id=?", (req.class_id, user_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="用户不存在")
    return {"ok": True, "user_id": user_id, "new_class_id": req.class_id}


class TeacherClassesRequest(BaseModel):
    class_ids: list


@router.put("/users/{user_id}/teacher-classes")
def update_teacher_classes(user_id: str, req: TeacherClassesRequest, request: Request):
    """为教师分配多个班级（仅管理员）。"""
    _require_admin(request)
    with get_conn() as conn:
        row = conn.execute("SELECT role FROM users WHERE user_id=?", (user_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    if dict(row)["role"] != "teacher":
        raise HTTPException(status_code=400, detail="该用户不是教师")
    clean_ids = [c.strip() for c in req.class_ids if c and c.strip()]
    set_teacher_class_ids(user_id, clean_ids)
    return {"ok": True, "user_id": user_id, "class_ids": clean_ids}


@router.get("/users/{user_id}/teacher-classes")
def get_teacher_classes_api(user_id: str, request: Request):
    """获取教师的班级列表（仅管理员）。"""
    _require_admin(request)
    return {"class_ids": get_teacher_class_ids(user_id)}


@router.get("/classes")
def list_classes(request: Request):
    """获取所有学生班级列表（仅管理员，排除教师 JSON 数组）。"""
    _require_admin(request)
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT DISTINCT class_id FROM users WHERE role='student' AND class_id IS NOT NULL AND class_id != '' ORDER BY class_id"
        ).fetchall()
    return {"classes": [r[0] for r in rows]}


@router.delete("/users/{user_id}")
def delete_user(user_id: str, request: Request):
    """删除用户（仅管理员）。"""
    admin = _require_admin(request)
    if admin["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="不能删除自己")
    with get_conn() as conn:
        conn.execute("DELETE FROM auth_tokens WHERE user_id=?", (user_id,))
        result = conn.execute("DELETE FROM users WHERE user_id=?", (user_id,))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="用户不存在")
    return {"ok": True, "deleted_user_id": user_id}


# ── 2. 全局数据监控 ──────────────────────────────────────────────

@router.get("/dashboard")
def admin_dashboard(request: Request):
    """
    全局大盘视图：展示全校项目总数、总体健康度/平均分、
    以及高频触发的 Top 3 业务漏洞。
    """
    _require_admin(request)

    all_projects = get_all_projects()
    total_projects = len(all_projects)

    # 计算平均分
    total_score = 0
    scored_count = 0
    all_diagnoses: list[str] = []
    class_stats: dict[str, dict] = {}

    for proj in all_projects:
        scores = proj.get("scores", {})
        if scores:
            avg = sum(scores.values()) / len(scores) if scores else 0
            total_score += avg
            scored_count += 1

        # 收集诊断信息
        diagnosis = proj.get("diagnosis", [])
        if isinstance(diagnosis, str):
            try:
                diagnosis = json.loads(diagnosis)
            except (json.JSONDecodeError, TypeError):
                diagnosis = [diagnosis] if diagnosis else []
        all_diagnoses.extend(diagnosis)

        # 按班级统计项目（只计入 student 的项目）
        owner_id = proj.get("owner_id", "unknown")
        with get_conn() as conn:
            user_row = conn.execute(
                "SELECT class_id, role FROM users WHERE user_id=?", (owner_id,)
            ).fetchone()
        if not user_row:
            continue
        user_data = dict(user_row)
        if user_data.get("role") != "student":
            continue
        class_id = user_data.get("class_id", "") or "未分班"

        if class_id not in class_stats:
            class_stats[class_id] = {"project_count": 0, "students": set(), "teachers": set()}
        class_stats[class_id]["project_count"] += 1
        class_stats[class_id]["students"].add(owner_id)

    # 统计每个班级的教师
    with get_conn() as conn:
        teacher_rows = conn.execute(
            "SELECT user_id, class_id FROM users WHERE role='teacher' AND class_id IS NOT NULL AND class_id != ''"
        ).fetchall()
    for tr in teacher_rows:
        t_class = tr[1]
        if t_class not in class_stats:
            class_stats[t_class] = {"project_count": 0, "students": set(), "teachers": set()}
        class_stats[t_class]["teachers"].add(tr[0])

    avg_score = round(total_score / scored_count, 2) if scored_count else 0

    # 高频业务漏洞 Top N
    vuln_freq: dict[str, int] = {}
    for d in all_diagnoses:
        vuln_freq[d] = vuln_freq.get(d, 0) + 1
    top_vulns = sorted(vuln_freq.items(), key=lambda x: x[1], reverse=True)[:5]

    # 用户统计
    with get_conn() as conn:
        user_counts = {}
        for role in ("student", "teacher", "admin"):
            count = conn.execute(
                "SELECT COUNT(*) FROM users WHERE role=?", (role,)
            ).fetchone()[0]
            user_counts[role] = count

    # 转换 class_stats 中的 set 为 count
    class_summary = {
        k: {
            "project_count": v["project_count"],
            "student_count": len(v["students"]),
            "teacher_count": len(v.get("teachers", set())),
        }
        for k, v in class_stats.items()
    }

    # 评分分布 (0-2, 2-4, 4-6, 6-8, 8-10)
    score_distribution = {"0-2": 0, "2-4": 0, "4-6": 0, "6-8": 0, "8-10": 0}
    for proj in all_projects:
        scores = proj.get("scores", {})
        if scores:
            avg = sum(scores.values()) / len(scores)
            if avg < 2: score_distribution["0-2"] += 1
            elif avg < 4: score_distribution["2-4"] += 1
            elif avg < 6: score_distribution["4-6"] += 1
            elif avg < 8: score_distribution["6-8"] += 1
            else: score_distribution["8-10"] += 1

    # 阶段分布
    stage_distribution: dict[str, int] = {}
    for proj in all_projects:
        stage = proj.get("stage", "discovery") or "discovery"
        stage_distribution[stage] = stage_distribution.get(stage, 0) + 1

    # 班级平均分对比
    class_scores: dict[str, list[float]] = {}
    for proj in all_projects:
        scores = proj.get("scores", {})
        if not scores:
            continue
        avg = sum(scores.values()) / len(scores)
        owner_id = proj.get("owner_id", "")
        with get_conn() as conn:
            row = conn.execute("SELECT class_id FROM users WHERE user_id=?", (owner_id,)).fetchone()
        cls = (dict(row).get("class_id", "") if row else "") or "未分班"
        class_scores.setdefault(cls, []).append(avg)

    class_avg_scores = {
        k: round(sum(v) / len(v), 2) for k, v in class_scores.items() if v
    }

    return {
        "total_projects": total_projects,
        "average_score": avg_score,
        "scored_projects": scored_count,
        "top_vulnerabilities": [
            {"vulnerability": v, "count": c} for v, c in top_vulns
        ],
        "user_counts": user_counts,
        "class_summary": class_summary,
        "health_status": "healthy" if avg_score >= 5 else "at_risk" if avg_score >= 3 else "critical",
        "score_distribution": score_distribution,
        "stage_distribution": stage_distribution,
        "class_avg_scores": class_avg_scores,
    }


@router.get("/projects")
def admin_list_all_projects(request: Request):
    """管理员查看所有项目。"""
    _require_admin(request)
    projects = get_all_projects()
    # Enrich with owner info
    with get_conn() as conn:
        for p in projects:
            owner = conn.execute(
                "SELECT username, display_name, class_id FROM users WHERE user_id=?",
                (p.get("owner_id", ""),)
            ).fetchone()
            if owner:
                p["owner_name"] = dict(owner).get("display_name") or dict(owner).get("username", "")
                p["owner_class"] = dict(owner).get("class_id", "")
    return {"projects": projects, "total": len(projects)}


# ── 3. 批量操作 ──────────────────────────────────────────────────

class BatchClassRequest(BaseModel):
    user_ids: list[str]
    class_id: str


@router.put("/users/batch-class")
def batch_update_class(req: BatchClassRequest, request: Request):
    """批量分配班级（仅管理员）。"""
    _require_admin(request)
    with get_conn() as conn:
        for uid in req.user_ids:
            conn.execute("UPDATE users SET class_id=? WHERE user_id=?", (req.class_id, uid))
    return {"ok": True, "updated": len(req.user_ids)}


# ── 3b. 批量导入用户（CSV） ────────────────────────────────────────

@router.post("/users/batch-import")
async def batch_import_users(request: Request):
    """
    批量导入用户（仅管理员）。
    接受 JSON body: { "users": [ {"username": "...", "password": "...", "role": "student", "class_id": "...", "display_name": "..."}, ... ] }
    """
    _require_admin(request)
    body = await request.json()
    users_data = body.get("users", [])
    if not users_data:
        raise HTTPException(status_code=400, detail="用户列表为空")

    from services.database import save_user
    import uuid
    results = {"success": 0, "failed": 0, "errors": []}
    for i, u in enumerate(users_data):
        username = u.get("username", "").strip()
        password = u.get("password", "").strip()
        role = u.get("role", "student").strip()
        class_id = u.get("class_id", "").strip()
        display_name = u.get("display_name", "").strip()

        if not username or not password:
            results["failed"] += 1
            results["errors"].append(f"第{i+1}行: 用户名或密码为空")
            continue
        if role not in ("student", "teacher", "admin"):
            role = "student"

        user_id = f"user_{uuid.uuid4().hex[:8]}"
        ok = save_user(user_id, username, password, role, display_name or username, class_id)
        if ok:
            results["success"] += 1
        else:
            results["failed"] += 1
            results["errors"].append(f"第{i+1}行: 用户名 '{username}' 已存在")

    return results


# ── 3c. 批量删除用户 ──────────────────────────────────────────────

class BatchDeleteRequest(BaseModel):
    user_ids: list[str]


@router.delete("/users/batch-delete")
def batch_delete_users(req: BatchDeleteRequest, request: Request):
    """批量删除用户（仅管理员，不可删除自己）。"""
    admin = _require_admin(request)
    admin_id = admin["user_id"]
    if admin_id in req.user_ids:
        raise HTTPException(status_code=400, detail="不可删除自己")
    with get_conn() as conn:
        deleted = 0
        for uid in req.user_ids:
            cursor = conn.execute("DELETE FROM users WHERE user_id=?", (uid,))
            deleted += cursor.rowcount
    return {"ok": True, "deleted": deleted}


# ── 4. 数据导出 ──────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(request: Request):
    """导出全校数据为 Excel（仅管理员）。"""
    _require_admin(request)
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Users
    ws_users = wb.active
    ws_users.title = "用户列表"
    ws_users.append(["用户ID", "用户名", "显示名称", "角色", "班级", "注册时间"])
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT user_id, username, display_name, role, class_id, created_at FROM users ORDER BY created_at DESC"
        ).fetchall()
    for r in rows:
        ws_users.append([r["user_id"], r["username"], r["display_name"], r["role"], r["class_id"] or "", r["created_at"] or ""])

    # Sheet 2: Projects
    ws_proj = wb.create_sheet("项目列表")
    ws_proj.append(["项目ID", "项目名称", "行业", "所有者ID", "阶段", "创建时间"])
    projects = get_all_projects()
    for p in projects:
        ws_proj.append([
            p.get("project_id", ""), p.get("name", ""), p.get("industry", ""),
            p.get("owner_id", ""), p.get("stage", ""), p.get("created_at", ""),
        ])

    # Sheet 3: Scores
    ws_scores = wb.create_sheet("项目评分")
    ws_scores.append(["项目名称", "行业", "empathy", "ideation", "business", "execution", "pitching", "诊断结果"])
    for p in projects:
        scores = p.get("scores", {})
        ws_scores.append([
            p.get("name", ""), p.get("industry", ""),
            scores.get("empathy", ""), scores.get("ideation", ""),
            scores.get("business", ""), scores.get("execution", ""),
            scores.get("pitching", ""),
            ", ".join(p.get("diagnosis", [])) if isinstance(p.get("diagnosis"), list) else str(p.get("diagnosis", "")),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ventureai_admin_export.xlsx"},
    )
