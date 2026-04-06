import io
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from services.session_store import get_all_projects, get_project
from services.database import (
    get_sessions_for_project, get_chat_history,
    set_competition_date as db_set_competition_date,
    save_annotation, get_annotations_for_project,
    get_all_ratings,
    save_instructor_review_notes, get_instructor_review_notes,
    save_teacher_intervention, get_teacher_interventions, delete_teacher_intervention,
)
from services.evidence_tracer import EvidenceTracer
from collections import Counter

router = APIRouter(prefix="/api/teacher", tags=["teacher"])

DIAGNOSIS_RUBRIC_MAP = {
    "需求真实性待验证": "R1",
    "需要补充用户访谈数据": "R2",
    "竞品认知不足": "R5",
    "竞品认知需深化": "R5",
    "护城河不清晰": "R7",
    "盈利模式需细化": "R4,R6",
    "商业模式未闭环": "R4",
    "商业模式完全缺失": "R4",
    "技术路线需评估": "R3",
    "目标用户太宽泛": "R1",
}

HIGH_RISK_THRESHOLD = 5.0


@router.get("/dashboard")
def get_dashboard(class_id: str = "class_2026_spring"):
    projects = get_all_projects()

    if not projects:
        return {
            "class_id": class_id,
            "total_students": 0,
            "total_projects": 0,
            "avg_scores": {"empathy": 0, "ideation": 0, "business": 0, "execution": 0, "pitching": 0},
            "top_mistakes": [],
            "high_risk_projects": [],
            "suggestions": ["暂无数据，等待学生创建项目并与AI教练对话后将自动生成分析"],
        }

    owner_ids = set(p.get("owner_id", "") for p in projects)

    dims = ["empathy", "ideation", "business", "execution", "pitching"]
    scored_projects = [p for p in projects if p.get("scores") and any(p["scores"].get(d, 0) > 0 for d in dims)]

    if scored_projects:
        avg_scores = {}
        for d in dims:
            vals = [p["scores"].get(d, 0) for p in scored_projects]
            avg_scores[d] = round(sum(vals) / len(vals), 1)
    else:
        avg_scores = {d: 0 for d in dims}

    all_diagnosis = []
    for p in projects:
        all_diagnosis.extend(p.get("diagnosis", []))

    mistake_counter = Counter(all_diagnosis)
    total_count = max(len(projects), 1)
    top_mistakes = []
    for rank, (mistake, count) in enumerate(mistake_counter.most_common(5), 1):
        freq = f"{round(count / total_count * 100)}%"
        rubric = DIAGNOSIS_RUBRIC_MAP.get(mistake, "")
        top_mistakes.append({"rank": rank, "mistake": mistake, "frequency": freq, "affected_rubric": rubric})

    high_risk = []
    for p in scored_projects:
        vals = [p["scores"].get(d, 0) for d in dims]
        avg = round(sum(vals) / len(vals), 1) if vals else 0
        if avg < HIGH_RISK_THRESHOLD:
            diag = p.get("diagnosis", ["综合评分偏低"])
            high_risk.append({
                "project_id": p["project_id"],
                "name": p["name"],
                "risk": diag[0] if diag else "综合评分偏低",
                "score": avg,
            })
    high_risk.sort(key=lambda x: x["score"])

    suggestions = _generate_suggestions(avg_scores, top_mistakes, high_risk)

    # A6-2: average_rubric_score — computed from rubric_scores stored in projects
    rubric_totals: dict[str, list[float]] = {}
    for p in projects:
        rs = p.get("rubric_scores") or {}
        for k, v in rs.items():
            rubric_totals.setdefault(k, []).append(float(v))
    average_rubric_score = {
        k: round(sum(v) / len(v), 2) for k, v in rubric_totals.items()
    } if rubric_totals else {}

    # A6-2: rule_trigger_frequency — count H-rule triggers across all sessions
    from services.database import get_conn
    rule_freq: dict[str, int] = {}
    try:
        with get_conn() as conn:
            rows = conn.execute(
                "SELECT content FROM chat_messages WHERE role='assistant'"
            ).fetchall()
        import re as _re
        for row in rows:
            for m in _re.finditer(r"\b(H\d{1,2})\b", row["content"]):
                rule_id = m.group(1)
                rule_freq[rule_id] = rule_freq.get(rule_id, 0) + 1
    except Exception:
        pass

    return {
        "class_id": class_id,
        "total_students": len(owner_ids),
        "total_projects": len(projects),
        "avg_scores": avg_scores,
        "top_mistakes": top_mistakes,
        "high_risk_projects": high_risk,
        "suggestions": suggestions,
        "average_rubric_score": average_rubric_score,
        "rule_trigger_frequency": rule_freq,
    }


def _generate_suggestions(avg_scores: dict, top_mistakes: list, high_risk: list) -> list:
    suggestions = []
    dim_names = {"empathy": "痛点发现", "ideation": "方案策划", "business": "商业建模", "execution": "资源杠杆", "pitching": "路演表达"}
    dim_topics = {
        "empathy": "用户访谈与需求验证方法",
        "ideation": "方案设计与技术可行性评估",
        "business": "商业模式画布与盈利逻辑（护城河理论）",
        "execution": "里程碑规划与资源获取策略",
        "pitching": "路演叙事结构与数据支撑技巧",
    }

    if avg_scores:
        weakest = min((k for k in avg_scores if avg_scores[k] > 0), key=lambda k: avg_scores[k], default=None)
        if weakest and avg_scores[weakest] < 6:
            suggestions.append(
                f"班级在「{dim_names[weakest]}」维度最薄弱（平均 {avg_scores[weakest]} 分），"
                f"建议下周重点讲解{dim_topics[weakest]}"
            )

    if top_mistakes:
        top = top_mistakes[0]
        suggestions.append(f"{top['frequency']}的团队存在「{top['mistake']}」问题，建议安排专项练习或工作坊")

    if high_risk:
        names = "、".join(p["name"] for p in high_risk[:3])
        suggestions.append(f"建议对以下高风险项目安排1对1辅导：{names}")

    if not suggestions:
        suggestions.append("班级整体表现良好，建议继续推进当前教学计划")

    return suggestions


@router.get("/projects")
def get_all_class_projects(class_id: str = "class_2026_spring"):
    projects = get_all_projects()
    return {"projects": projects}


@router.get("/project/{project_id}/review")
def get_project_review(project_id: str):
    proj = get_project(project_id)
    if not proj:
        return {"project_id": project_id, "error": "项目不存在"}

    scores = proj.get("scores", {})
    diagnosis = proj.get("diagnosis", [])
    rubric_scores = _scores_to_rubric(scores, diagnosis)
    total = sum(item["score"] for item in rubric_scores.values())

    evidence = []
    if proj.get("description"):
        evidence.append({"rubric": "R1", "quote": f"\"{proj['description']}\"", "source": "项目描述"})
    if proj.get("industry"):
        evidence.append({"rubric": "R5", "quote": f"行业：{proj['industry']}", "source": "项目信息"})

    suggestions = []
    for d in diagnosis[:3]:
        if "访谈" in d or "证据" in d:
            suggestions.append(f"针对「{d}」：完成5人用户访谈，整理关键引用")
        elif "商业" in d or "盈利" in d:
            suggestions.append(f"针对「{d}」：完成 Lean Canvas，明确收入-成本结构")
        elif "竞品" in d or "护城河" in d:
            suggestions.append(f"针对「{d}」：做竞品对比矩阵（至少3个替代方案）")
        else:
            suggestions.append(f"针对「{d}」：补充相关分析和证据材料")

    if not suggestions:
        suggestions.append("继续深化当前项目各维度，建议与AI教练进行更多对话以获得精准诊断")

    # Enrich evidence trace using EvidenceTracer across all project sessions
    try:
        sessions = get_sessions_for_project(project_id)
        tracer = EvidenceTracer(project_id)
        all_messages: list[dict] = []
        for sess in sessions:
            msgs = get_chat_history(sess["session_id"])
            all_messages.extend(msgs)
        if all_messages:
            tracer.ingest(all_messages)
            ev_summary = tracer.summarize()
            for ev in ev_summary.get("evidence_list", [])[:10]:
                rubric_tag = ev["rubric_tags"][0] if ev["rubric_tags"] else "general"
                rubric_short = rubric_tag.split("_")[0]
                evidence.append({
                    "rubric": rubric_short,
                    "quote": f"「{ev['text'][:80]}」",
                    "source": f"第{ev['turn']}轮对话（{ev['ev_type']}）",
                })
    except Exception:
        pass

    # A6-1: instructor_review_notes
    review_notes_row = get_instructor_review_notes(project_id)
    instructor_review_notes = review_notes_row["notes"] if review_notes_row else {}

    return {
        "project_id": project_id,
        "rubric_scores": rubric_scores,
        "total_score": total,
        "max_score": len(rubric_scores) * 10,  # rubric here uses 5-dim scores (0-10 each)
        "evidence_trace": evidence,
        "revision_suggestions": suggestions,
        "instructor_review_notes": instructor_review_notes,
    }


@router.get("/project/{project_id}/conversations")
def get_project_conversations(project_id: str):
    """
    教师查看某项目的所有学生对话记录。
    Returns all chat sessions and messages for a given project.
    """
    proj = get_project(project_id)
    if not proj:
        return {"error": "项目不存在"}

    sessions = get_sessions_for_project(project_id)
    result = []
    for sess in sessions:
        msgs = get_chat_history(sess["session_id"])
        result.append({
            "session_id": sess["session_id"],
            "agent_type": sess.get("agent_type", "auto"),
            "created_at": sess.get("created_at", ""),
            "message_count": len(msgs),
            "messages": msgs,
        })

    # Build evidence summary from all messages
    ev_summary = None
    try:
        all_messages = []
        for sess_data in result:
            all_messages.extend(sess_data["messages"])
        if all_messages:
            tracer = EvidenceTracer(project_id)
            tracer.ingest(all_messages)
            ev_summary = tracer.summarize()
    except Exception:
        pass

    return {
        "project_id": project_id,
        "project_name": proj.get("name", ""),
        "owner_id": proj.get("owner_id", ""),
        "sessions": result,
        "evidence_summary": ev_summary,
    }


@router.get("/export/excel")
def export_excel(class_id: str = "class_2026_spring"):
    """Export all projects as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请运行 pip install openpyxl")

    projects = get_all_projects()
    dims = ["empathy", "ideation", "business", "execution", "pitching"]
    dim_labels = {"empathy": "痛点发现", "ideation": "方案策划", "business": "商业建模",
                  "execution": "资源杠杆", "pitching": "路演表达"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全班项目评估"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["项目名称", "行业", "当前阶段", "学生ID"] + [dim_labels[d] for d in dims] + ["综合平均", "诊断问题", "创建日期"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    stage_labels = {"discovery": "痛点发现", "ideation": "方案策划", "modeling": "商业建模",
                    "execution": "资源杠杆", "pitching": "路演表达"}
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, proj in enumerate(projects, 2):
        scores = proj.get("scores", {})
        vals = [scores.get(d, 0) for d in dims]
        avg = round(sum(vals) / len(vals), 1) if any(vals) else 0
        diag = "；".join(proj.get("diagnosis", []))

        row_data = [
            proj.get("name", ""),
            proj.get("industry", ""),
            stage_labels.get(proj.get("stage", ""), proj.get("stage", "")),
            proj.get("owner_id", ""),
        ] + vals + [avg, diag, proj.get("created_at", "")]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if avg < 5.0 and avg > 0:
                cell.fill = red_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # Summary sheet
    ws2 = wb.create_sheet("班级汇总")
    scored = [p for p in projects if p.get("scores") and any(p["scores"].get(d, 0) > 0 for d in dims)]
    ws2.append(["统计项", "数值"])
    ws2.append(["项目总数", len(projects)])
    ws2.append(["已评分项目", len(scored)])
    ws2.append(["高风险项目（<5分）", sum(1 for p in scored if sum(p["scores"].get(d, 0) for d in dims) / 5 < 5)])
    if scored:
        for d in dims:
            avg_d = round(sum(p["scores"].get(d, 0) for p in scored) / len(scored), 1)
            ws2.append([f"班级平均-{dim_labels[d]}", avg_d])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ventureai_class_report.xlsx"}
    )


def _scores_to_rubric(scores: dict, diagnosis: list) -> dict:
    e = scores.get("empathy", 5)
    i_score = scores.get("ideation", 5)
    b = scores.get("business", 5)
    x = scores.get("execution", 5)
    p = scores.get("pitching", 5)
    diag_text = " ".join(diagnosis)

    def missing_if(keywords: list, evidence_label: str) -> list:
        return [evidence_label] if any(kw in diag_text for kw in keywords) else []

    return {
        "R1": {
            "score": e,
            "justification": "痛点定义" + ("清晰" if e >= 7 else "待加强" if e >= 5 else "严重不足"),
            "missing": missing_if(["需求", "痛点", "用户太宽泛"], "用户访谈记录"),
        },
        "R2": {
            "score": max(1, e - 2),
            "justification": "用户证据" + ("充分" if e >= 8 else "不足" if e >= 5 else "严重缺乏"),
            "missing": missing_if(["访谈", "证据"], "访谈引用与行为数据") or (["用户访谈记录"] if e < 7 else []),
        },
        "R3": {
            "score": i_score,
            "justification": "技术方案" + ("可行" if i_score >= 7 else "基本可行" if i_score >= 5 else "存疑"),
            "missing": missing_if(["技术路线"], "技术路线图"),
        },
        "R4": {
            "score": b,
            "justification": "商业模式" + ("闭环" if b >= 7 else "待完善" if b >= 5 else "缺失"),
            "missing": missing_if(["商业", "盈利"], "商业模式画布") or (["商业模式画布"] if b < 6 else []),
        },
        "R5": {
            "score": max(1, round((e + b) / 2)),
            "justification": "市场竞争分析" + ("合理" if b >= 7 else "需完善"),
            "missing": missing_if(["竞品", "护城河"], "竞品对比表"),
        },
        "R6": {
            "score": max(1, b - 1),
            "justification": "财务逻辑" + ("合理" if b >= 7 else "待验证" if b >= 5 else "缺失"),
            "missing": (["CAC/LTV计算"] if b < 6 else []),
        },
        "R7": {
            "score": i_score,
            "justification": "创新差异化" + ("明确" if i_score >= 7 else "待量化" if i_score >= 5 else "不清晰"),
            "missing": missing_if(["创新"], "对比矩阵"),
        },
        "R8": {
            "score": x,
            "justification": "团队执行力" + ("强" if x >= 7 else "中等" if x >= 5 else "薄弱"),
            "missing": missing_if(["里程碑", "团队"], "里程碑计划"),
        },
        "R9": {
            "score": p,
            "justification": "表达与材料" + ("优秀" if p >= 7 else "一般" if p >= 5 else "需加强"),
            "missing": missing_if(["叙事", "表达"], "路演PPT"),
        },
        "R10": {
            "score": max(1, round((b + x) / 2 - 1)),
            "justification": "合规与社会责任" + ("已评估" if b >= 7 else "需补充"),
            "missing": missing_if(["合规", "数据", "隐私"], "合规清单与知识产权说明"),
        },
        "R11": {
            "score": max(1, round((i_score + x) / 2 - 1)),
            "justification": "增长与规模化" + ("路径清晰" if i_score >= 7 else "规划不足"),
            "missing": missing_if(["规模", "增长"], "阶段性增长路线图"),
        },
    }


# ── F2: Competition Date ─────────────────────────────────────────────

class CompetitionDateRequest(BaseModel):
    competition_date: str  # ISO format: "2026-04-15"


@router.post("/project/{project_id}/competition-date")
def set_comp_date(project_id: str, req: CompetitionDateRequest):
    db_set_competition_date(project_id, req.competition_date)
    return {"ok": True, "project_id": project_id, "competition_date": req.competition_date}


# ── F3: Teacher Annotations ──────────────────────────────────────────

class AnnotationRequest(BaseModel):
    project_id: str
    session_id: str
    message_idx: int
    teacher_id: str = "teacher"
    action: str       # 'like' | 'dislike' | 'note'
    note_text: str = ""


@router.post("/annotation")
def create_annotation(req: AnnotationRequest):
    save_annotation(req.project_id, req.session_id, req.message_idx,
                    req.teacher_id, req.action, req.note_text)
    return {"ok": True}


@router.get("/project/{project_id}/annotations")
def get_annotations(project_id: str):
    return {"annotations": get_annotations_for_project(project_id)}


# ── F5: Session Ratings (teacher view) ──────────────────────────────

@router.get("/ratings")
def get_all_session_ratings():
    ratings = get_all_ratings()
    avg = round(sum(r["rating"] for r in ratings) / len(ratings), 1) if ratings else 0
    return {"ratings": ratings, "avg_rating": avg, "total": len(ratings)}


# ── 知识覆盖汇总 ────────────────────────────────────────────────────

# All tutor concepts that can be taught
ALL_CONCEPTS = [
    "PMF", "CAC", "LTV", "TAM", "SAM", "SOM", "JTBD", "AARRR",
    "SWOT", "BEP", "Lean Canvas", "精益画布", "商业模式画布",
    "波特五力", "价值主张", "护城河", "用户画像", "竞品分析",
    "TAM/SAM/SOM", "产品市场契合", "获客成本", "终身价值",
    "市场规模", "商业模式", "精益创业", "最小可行产品",
    "PMF验证", "联合创始人", "股权结构", "投资人关系",
]

# Concept → explanation keywords that appear in assistant messages
CONCEPT_MENTION_PATTERNS = {
    "PMF": ["PMF", "产品市场契合", "product market fit"],
    "CAC": ["CAC", "获客成本", "customer acquisition cost"],
    "LTV": ["LTV", "终身价值", "lifetime value"],
    "TAM": ["TAM", "总体市场规模", "可寻址市场"],
    "SAM": ["SAM", "可服务市场"],
    "SOM": ["SOM", "可获取市场"],
    "JTBD": ["JTBD", "jobs to be done", "用户任务理论"],
    "AARRR": ["AARRR", "增长黑客", "海盗指标"],
    "SWOT": ["SWOT", "优势劣势机会威胁"],
    "BEP": ["BEP", "盈亏平衡", "盈亏平衡点"],
    "精益画布": ["精益画布", "lean canvas"],
    "商业模式画布": ["商业模式画布", "BMC", "business model canvas"],
    "波特五力": ["波特五力", "porter", "五力模型"],
    "价值主张": ["价值主张", "value proposition"],
    "护城河": ["护城河", "moat", "竞争壁垒"],
    "用户画像": ["用户画像", "persona", "用户画像分析"],
    "竞品分析": ["竞品分析", "竞争对手分析", "竞争格局"],
    "最小可行产品": ["MVP", "最小可行产品", "minimum viable product"],
    "股权结构": ["股权", "股权结构", "cap table"],
    "精益创业": ["精益创业", "lean startup"],
}


# ── A6-1: Instructor Review Notes ───────────────────────────────────

class ReviewNotesRequest(BaseModel):
    teacher_id: str = "teacher"
    notes: dict  # {R1: "...", overall: "...", ...}


@router.post("/project/{project_id}/review-notes")
def save_review_notes(project_id: str, req: ReviewNotesRequest):
    """Save instructor review notes for a project."""
    save_instructor_review_notes(project_id, req.teacher_id, req.notes)
    return {"ok": True, "project_id": project_id}


# ── A6-3: Teacher Interventions ──────────────────────────────────────

class InterventionRequest(BaseModel):
    teacher_id: str = "teacher"
    project_id: str
    trigger_keyword: str  # keyword in student message that triggers this instruction
    instruction: str      # instruction injected into coach system prompt
    priority: int = 1     # higher = injected first


@router.post("/interventions")
def create_intervention(req: InterventionRequest):
    """Save a teacher intervention rule."""
    import uuid as _uuid
    intervention_id = str(_uuid.uuid4())
    save_teacher_intervention(
        intervention_id, req.teacher_id, req.project_id,
        req.trigger_keyword, req.instruction, req.priority
    )
    return {"ok": True, "intervention_id": intervention_id}


@router.get("/interventions/{project_id}")
def list_interventions(project_id: str):
    """List all active intervention rules for a project."""
    items = get_teacher_interventions(project_id)
    return {"interventions": items}


@router.delete("/interventions/{intervention_id}")
def remove_intervention(intervention_id: str):
    """Delete an intervention rule."""
    delete_teacher_intervention(intervention_id)
    return {"ok": True}


# ── A6-4: Capability Profile ─────────────────────────────────────────

_CAPABILITY_SYSTEM = """你是一位创新创业教育分析专家。
请根据学生在以下对话记录中的表现，生成一份五维能力画像评分报告。

五个维度（每项 0-10 分，10为最高）：
1. 同理心（Empathy）：识别用户真实痛点、站在用户角度思考的能力
2. 创意发散（Ideation）：提出有创意、有价值的解决方案的能力
3. 商业建模（Business）：理解商业逻辑、收入模型、市场策略的能力
4. 执行规划（Execution）：制定可行计划、分配资源、管理里程碑的能力
5. 逻辑表达（Logic）：清晰论证、数据支撑、结构化表达的能力

请严格按以下 JSON 格式输出：
<!--CAPABILITY_PROFILE:
{
  "empathy": {"score": 7, "evidence": "学生提到...", "suggestion": "建议..."},
  "ideation": {"score": 6, "evidence": "...", "suggestion": "..."},
  "business": {"score": 5, "evidence": "...", "suggestion": "..."},
  "execution": {"score": 4, "evidence": "...", "suggestion": "..."},
  "logic": {"score": 6, "evidence": "...", "suggestion": "..."},
  "overall_comment": "综合评价..."
}
-->

在 JSON 之前，给出 1-2 段整体能力评价。"""


@router.get("/project/{project_id}/capability-profile")
def get_capability_profile(project_id: str):
    """
    A6-4: 从项目对话历史生成五维能力画像。
    最多取最近 3 轮会话 (window) 进行分析。
    """
    import re as _re
    import json as _json
    from config import USE_MOCK_API
    from services.claude_client import chat_completion

    proj = get_project(project_id)
    if not proj:
        return {"error": "项目不存在"}

    # Gather recent conversation messages (up to 3 sessions)
    sessions = get_sessions_for_project(project_id)
    all_msgs: list[dict] = []
    for sess in sessions[-3:]:
        all_msgs.extend(get_chat_history(sess["session_id"]))

    if not all_msgs:
        return {
            "project_id": project_id,
            "capability_profile": None,
            "message": "暂无对话记录，请先与AI教练进行对话",
        }

    if USE_MOCK_API:
        profile = {
            "empathy": {"score": 7, "evidence": "学生能描述具体用户场景", "suggestion": "尝试量化痛点频率"},
            "ideation": {"score": 6, "evidence": "方案有一定创意", "suggestion": "拓展差异化方向"},
            "business": {"score": 5, "evidence": "商业模式基本清晰", "suggestion": "补充CAC/LTV分析"},
            "execution": {"score": 5, "evidence": "有初步时间线规划", "suggestion": "细化里程碑指标"},
            "logic": {"score": 6, "evidence": "论述结构较清晰", "suggestion": "增加数据支撑"},
            "overall_comment": "学生展现出较好的同理心和表达能力，商业和执行维度有较大提升空间。",
        }
        return {"project_id": project_id, "capability_profile": profile, "session_count": len(sessions)}

    raw = chat_completion(_CAPABILITY_SYSTEM, all_msgs)

    from services.marker_parser import parse_capability_profile, clean_reply
    profile = parse_capability_profile(raw)
    clean = clean_reply(raw)

    return {
        "project_id": project_id,
        "capability_profile": profile,
        "summary": clean,
        "session_count": len(sessions),
    }


@router.get("/knowledge-coverage")
def get_knowledge_coverage():
    """
    扫描所有 tutor 会话，统计哪些概念已被讲解、各项目学生学习情况。
    """
    from services.database import get_conn

    with get_conn() as conn:
        # Get all tutor/hybrid sessions
        sessions = conn.execute(
            "SELECT cs.session_id, cs.project_id, p.owner_id, p.name as project_name "
            "FROM chat_sessions cs LEFT JOIN projects p ON cs.project_id = p.project_id "
            "WHERE cs.agent_type IN ('tutor', 'hybrid', 'auto') ORDER BY cs.created_at"
        ).fetchall()

    # concept → {count: int, projects: set, students: set}
    concept_stats: dict[str, dict] = {
        c: {"count": 0, "projects": set(), "students": set()} for c in CONCEPT_MENTION_PATTERNS
    }

    # project → covered concepts
    project_coverage: dict[str, set] = {}

    for sess in sessions:
        sid = sess["session_id"]
        pid = sess["project_id"] or ""
        student = sess["owner_id"] or ""

        with get_conn() as conn:
            msgs = conn.execute(
                "SELECT role, content FROM chat_messages WHERE session_id=? ORDER BY id",
                (sid,)
            ).fetchall()

        # Only scan assistant messages (AI taught concepts)
        full_text = " ".join(m["content"] for m in msgs if m["role"] == "assistant")

        for concept, patterns in CONCEPT_MENTION_PATTERNS.items():
            if any(p.lower() in full_text.lower() for p in patterns):
                concept_stats[concept]["count"] += 1
                if pid:
                    concept_stats[concept]["projects"].add(pid)
                    if pid not in project_coverage:
                        project_coverage[pid] = set()
                    project_coverage[pid].add(concept)
                if student:
                    concept_stats[concept]["students"].add(student)

    # Build coverage summary
    covered = [c for c, s in concept_stats.items() if s["count"] > 0]
    not_covered = [c for c, s in concept_stats.items() if s["count"] == 0]

    # Top concepts by frequency
    top_concepts = sorted(
        [{"concept": c, "count": s["count"], "student_count": len(s["students"]), "project_count": len(s["projects"])}
         for c, s in concept_stats.items() if s["count"] > 0],
        key=lambda x: x["count"], reverse=True
    )

    # Per-project coverage
    proj_coverage_list = []
    for pid, concepts in project_coverage.items():
        proj_coverage_list.append({
            "project_id": pid,
            "covered_concepts": sorted(concepts),
            "coverage_rate": round(len(concepts) / len(CONCEPT_MENTION_PATTERNS) * 100, 1),
        })

    return {
        "total_concepts": len(CONCEPT_MENTION_PATTERNS),
        "covered_count": len(covered),
        "not_covered_count": len(not_covered),
        "coverage_rate": round(len(covered) / len(CONCEPT_MENTION_PATTERNS) * 100, 1),
        "covered_concepts": covered,
        "not_covered_concepts": not_covered,
        "top_concepts": top_concepts[:15],
        "project_coverage": proj_coverage_list,
    }
